import openpyxl
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
import pandas as pd
import glob
import os

def formattingExcel(eachFolder):
    file = glob.glob(eachFolder+"/*")[0]
    excelFile = openpyxl.load_workbook(file)
    
    greenFill = PatternFill(start_color='FFC000',end_color='FFC000',fill_type='solid')
    thinBorder = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for sheetName in excelFile.sheetnames:
        sheet = excelFile[sheetName]
        sheet.delete_rows(1, 1)

        maxRow = sheet.max_row
        maxColumn = sheet.max_column
        for row in sheet.iter_rows(min_row=1, max_row=maxRow, min_col=1, max_col=maxColumn):
            for cell in row:
                cell.border = thinBorder
        
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column_letter].width = max_length + 2
    excelFile.save(file)
    

def mergeExcels(outFolder):

    allFolders = glob.glob(outFolder+"/*")
    
    for eachFolder in allFolders:
        fileDict = {}
        allFiles = glob.glob(eachFolder+"/*")
        outputFile = os.path.dirname(allFiles[0]) + "/" + os.path.basename(allFiles[0].split("_page")[0]+".xlsx")
        
        for file in allFiles:
            sheetNames = pd.ExcelFile(file).sheet_names
            if len(sheetNames) > 1:
                for sheet in sheetNames:
                    fileDict.setdefault(file, []).append(os.path.basename(file).replace(".xlsx","").split("_")[1] + "_" + sheet)

            else:
                fileDict.setdefault(file, os.path.basename(file).replace(".xlsx","").split("_")[1])
        
        with pd.ExcelWriter(outputFile) as writer:
            for file, sheet_names in fileDict.items():
                if isinstance(sheet_names, str):
                    data = pd.read_excel(file)
                    data.to_excel(writer, sheet_name=sheet_names, index=False)
                else:
                    for sheet_name in sheet_names:
                        data = pd.read_excel(file, sheet_name=sheet_name.split("_")[1])
                        data.to_excel(writer, sheet_name=sheet_name, index=False)


        for eachFile in allFiles:
            os.remove(eachFile)
        
        formattingExcel(eachFolder)

        name = eachFolder.split('\\')[-1]
        fString = f"Excel Formatting completed for {name}"
        print(fString)

# mergeExcels('./Output')